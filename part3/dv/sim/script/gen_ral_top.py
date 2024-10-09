#----------------------------------------------------------------------------------
# This code is copyrighted by BrentWang and cannot be used for commercial purposes
# The github address:https://github.com/brentwang-lab/uvm_tb_gen                   
# You can refer to the book <UVM Experiment Guide> for learning, this is on this github
# If you have any questions, please contact email:brent_wang@foxmail.com          
#----------------------------------------------------------------------------------
#                                                                                  
# Author  : BrentWang                                                              
# Project : UVM study                                                              
# Date    : Sat Jan 26 06:05:52 WAT 2022                                           
#----------------------------------------------------------------------------------
#                                                                                  
# Description:                                                                     
#     File for gen_ral_top.py                                                       
#----------------------------------------------------------------------------------
#!/usr/bin/python3

import re
import openpyxl
import argparse
import os
import sys

ral_head_text  = ""
ral_end_text  = ""
reg_rand_text = ""
field_path_text = ""
reg_build_text = ""
field_rand_text = ""
field_build_text = ""
reg_class_text = ""
block_class_text = ""
block_rand_text = ""
block_build_text = ""
topblock_class_text = ""

def getparse():
    parser = argparse.ArgumentParser(description="input and output file about register model")
    parser.add_argument('-i','--input_reg_table',help='please input the reg_table excel file, the format is xxx.xlsx',required=True, default='NO_Reg_Excel_Tabel')
    parser.add_argument('-o','--output_regmdl_file',help='there is no define uvm register model name',required=True, default='NO_Uvm_Register_Model')
    args = parser.parse_args()
    return args
#If this cell is empty, go up to find the value
def null_up_valid(p_sheet,p_row,p_col):
    if(p_sheet.cell(p_row,p_col).data_type != 'n'):
        return p_sheet.cell(p_row,p_col).value
    else:
        return null_up_valid(p_sheet,p_row-1,p_col)
#Find column position based on value
def get_col_from_value(p_sheet,p_value):
    for row in range (1,p_sheet.max_row+1):
        for col in range (1,p_sheet.max_column+1):
            if(p_sheet.cell(row,col).value ==p_value):
                return col
def gen_ral_head_text():
    global ral_head_text 
    ral_head_text += "`ifndef RAL_TOP__SV\n"
    ral_head_text += "`define RAL_TOP__SV\n"
    ral_head_text += "package ral_top_pkg;\n"
    ral_head_text += "import uvm_pkg::*;\n"
    ral_head_text += "`include \"uvm_macros.svh\"\n"

def gen_ral_end_text():
    global ral_end_text 
    ral_end_text += "endpackage : ral_top_pkg\n"
    ral_end_text += "`endif //RAL_TOP__SV\n"

def gen_field_rand_text(fld_name):
    global field_rand_text 
    field_rand_text += "\trand uvm_reg_field "+fld_name+"_field;\n"

def gen_reg_rand_text(reg_name):
    global reg_rand_text 
    reg_rand_text +="\trand "+reg_name+"_register "+reg_name+"_reg;\n" 

def gen_block_rand_text(block_name):
    global block_rand_text
    block_rand_text += "\trand "+block_name+"_block "+block_name+"_blk;\n"

def gen_field_build_text(fld_name,size_value,bitstart_value,fld_type,rst_value):
    global field_build_text
    field_build_text += "\t\t"+fld_name+"_field = uvm_reg_field :: type_id :: create(\""+fld_name+"_field\");\n"
    field_build_text += "\t\t"+fld_name+"_field.configure(.parent(this),.size("+str(size_value)+"),.lsb_pos("+str(bitstart_value)+"),\n"
    field_build_text += "\t\t\t\t\t.access(\""+fld_type+"\"),.volatile(0),.reset("+str(rst_value)+"),\n"
    field_build_text += "\t\t\t\t\t.has_reset(1), .is_rand(1), .individually_accessible(0));\n"

def gen_reg_build_text(reg_name,regadr_value):
    global reg_build_text 
    global field_path_text
    reg_build_text +="\t\t"+reg_name+"_reg = "+reg_name+"_register :: type_id :: create(\""+reg_name+"_reg\");\n" 
    reg_build_text +="\t\t"+reg_name+"_reg.configure(this,null,\"\");\n"
    reg_build_text +="\t\t"+reg_name+"_reg.build();\n"
    reg_build_text +=field_path_text
    reg_build_text +="\t\tdefault_map.add_reg("+reg_name+"_reg,"+regadr_value+",\"RW\");\n"
    field_path_text = "\n"

def gen_block_build_text(block_name,blockaddr_value,blockhdlpath_value):
    global block_build_text
    block_build_text += "\t\t"+block_name+"_blk = "+block_name+"_block :: type_id :: create(\""+block_name+"_blk\",,get_full_name());\n"
    block_build_text += "\t\t"+block_name+"_blk.configure(this,\""+blockhdlpath_value+"\");\n"
    block_build_text += "\t\t"+block_name+"_blk.build();\n"
    block_build_text += "\t\tdefault_map.add_submap("+block_name+"_blk.default_map,"+blockaddr_value+");\n"
    block_build_text += "\n"

def gen_reg_class_text(reg_name):
    global reg_class_text
    global field_rand_text
    global field_build_text
    reg_class_text += "class "+reg_name+"_register extends uvm_reg;\n"
    reg_class_text += "\t`uvm_object_utils("+reg_name+"_register)\n"
    reg_class_text += field_rand_text;
    reg_class_text += "\tfunction new(string name = \""+reg_name+"_register\");\n"
    reg_class_text += "\t\tsuper.new(name,32,UVM_CVR_ALL);\n"
    reg_class_text += "\tendfunction : new\n"
    reg_class_text += "\tvirtual function void build();\n"
    reg_class_text += field_build_text;
    reg_class_text += "\tendfunction : build\n"
    reg_class_text += "endclass : "+reg_name+"_register\n"
    field_rand_text = "\n"
    field_build_text = "\n"

def gen_block_class_text(block_name):
    global block_class_text
    global reg_rand_text
    global reg_build_text
    block_class_text += "class "+block_name+"_block extends uvm_reg_block;\n"
    block_class_text += "\t`uvm_object_utils("+block_name+"_block)\n"
    block_class_text += reg_rand_text
    block_class_text += "\tfunction new(string name = \""+block_name+"_block\");\n";
    block_class_text += "\t\tsuper.new(name,UVM_CVR_ALL);\n";
    block_class_text += "\tendfunction : new\n";
    block_class_text += "\tfunction void build();\n";
    block_class_text += "\t\tdefault_map = create_map(\"default_map\",0,4,UVM_LITTLE_ENDIAN);\n";
    block_class_text += reg_build_text;
    block_class_text += "\tendfunction : build\n";
    block_class_text += "endclass: "+block_name+"_block\n";
    reg_rand_text = "\n"
    reg_build_text = "\n"

def gen_topblock_class_text():
    global topblock_class_text
    global block_rand_text
    global block_build_text
    topblock_class_text += "class ral_top_block extends uvm_reg_block;\n"
    topblock_class_text += "\t`uvm_object_utils(ral_top_block)\n"
    topblock_class_text += block_rand_text
    topblock_class_text += "\tfunction new(string name = \"ral_top_block\");\n"
    topblock_class_text += "\t\tsuper.new(name);\n"
    topblock_class_text += "\tendfunction : new\n"
    topblock_class_text += "\tfunction void build();\n"
    topblock_class_text += "\t\tdefault_map = create_map(\"default_map\",0,4,UVM_LITTLE_ENDIAN);\n";
    topblock_class_text +=  block_build_text
    topblock_class_text += "\tendfunction : build\n"
    topblock_class_text += "endclass : ral_top_block\n"
    block_rand_text = "\n"
    block_build_text = "\n"

def gen_field_path_text(reg_name,fieldhdlpath_value,bitstart_value,size_value):
    global field_path_text
    field_path_text +="\t\t"+reg_name+"_reg.add_hdl_path_slice(\""+fieldhdlpath_value+"\","+str(bitstart_value)+","+str(size_value)+");\n"


def gen_ral_top(p_sheet,ModuleName):
    global field_path_text
    global reg_build_text 
    global field_rand_text 
    global field_build_text
    global reg_class_text
    global block_class_text
    global block_rand_text
    global block_build_text
    global reg_rand_text 
    global topblock_class_text
    global ral_end_text

    blockname_col = get_col_from_value(p_sheet,'BlockName')
    blockname_value = p_sheet.cell(2,blockname_col).value

    blockaddr_col = get_col_from_value(p_sheet,'BlockAddress')

    width_col = get_col_from_value(p_sheet,'Width')
    width_value = int(p_sheet.cell(2,width_col).value)

    reg_col = get_col_from_value(p_sheet,'RegName')
    fieldhdlpath_col = get_col_from_value(p_sheet,'FieldHdlPath')
    blockhdlpath_col = get_col_from_value(p_sheet,'BlockHdlPath')
    fld_col = get_col_from_value(p_sheet,'FieldName')
    rst_col = get_col_from_value(p_sheet,'ResetValue')
    bitstart_col = get_col_from_value(p_sheet,'BitStart')
    bitend_col = get_col_from_value(p_sheet,'BitEnd')
    access_col  = get_col_from_value(p_sheet,'Access')
    regadr_col  = get_col_from_value(p_sheet,'RegAddress')

    if os.path.exists("%s"%(ModuleName)):         
        os.remove("%s"%(ModuleName))
    fo=open("%s"%(ModuleName),"a")
    gen_ral_head_text()
    fo.write(ral_head_text)
    last_reg_name = ""
    last_block_name = ""
    last_blockhdlpath_value = ""
    last_regadr_value = ""
    last_blockaddr_value = ""
    for row in range (2,p_sheet.max_row+1):
        block_name = null_up_valid(p_sheet,row,blockname_col)
        reg_name  = null_up_valid(p_sheet,row,reg_col)
        fld_name  = p_sheet.cell(row,fld_col).value.lower()
        rst_value = p_sheet.cell(row,rst_col).value.replace('0x','\'h').replace('0X','\'h')
        bitstart_value       = p_sheet.cell(row,bitstart_col).value
        bitend_value       = p_sheet.cell(row,bitend_col).value
        size_value = int(bitend_value) - int(bitstart_value)+1
        fieldhdlpath_value = p_sheet.cell(row,fieldhdlpath_col).value
        blockhdlpath_value = null_up_valid(p_sheet,row,blockhdlpath_col)
        fld_type  = p_sheet.cell(row,access_col).value
        regadr_value = null_up_valid(p_sheet,row,regadr_col).replace('0x','\'h').replace('0X','\'h')
        blockaddr_value = null_up_valid(p_sheet,row,blockaddr_col).replace('0x','\'h').replace('0X','\'h')
        #Last line
        if(row == p_sheet.max_row):
            gen_field_rand_text(fld_name)
            gen_field_build_text(fld_name,size_value,bitstart_value,fld_type,rst_value)
            gen_field_path_text(reg_name,fieldhdlpath_value,bitstart_value,size_value)

            last_reg_name = reg_name
            last_regadr_value = regadr_value
            last_blockaddr_value = blockaddr_value
            last_block_name = block_name
            last_blockhdlpath_value = blockhdlpath_value
            reg_name = "NONE"
            block_name = "NONE"
            blockhdlpath_value = "NONE"
        if(fld_name.lower() != 'reserved'):
            if(reg_name != last_reg_name and last_reg_name != ""):
                gen_reg_rand_text(last_reg_name)
                gen_reg_build_text(last_reg_name,last_regadr_value)
                gen_reg_class_text(last_reg_name)
                fo.write(reg_class_text)
                reg_class_text = "\n"
                if(block_name != last_block_name and last_block_name != ""):
                    gen_block_class_text(last_block_name)
                    fo.write(block_class_text)
                    block_class_text = "\n"
                    gen_block_rand_text(last_block_name)
                    gen_block_build_text(last_block_name,last_blockaddr_value,last_blockhdlpath_value)
            gen_field_rand_text(fld_name)
            gen_field_build_text(fld_name,size_value,bitstart_value,fld_type,rst_value)
            gen_field_path_text(reg_name,fieldhdlpath_value,bitstart_value,size_value)
            last_reg_name = reg_name
            last_block_name = block_name
            last_regadr_value = regadr_value
            last_blockaddr_value = blockaddr_value
            last_blockhdlpath_value = blockhdlpath_value
    gen_topblock_class_text()
    fo.write(topblock_class_text)
    topblock_class_text = "\n"

    gen_ral_end_text()
    fo.write(ral_end_text)
    ral_end_text  = "\n"
    fo.close()
    print("*****Successfully generated %s*****"%(ModuleName))
 

if __name__ == '__main__':
    args_t = getparse()
    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(args_t.input_reg_table, data_only=True)
    sheet = wb['REG_LIST']
    gen_ral_top(sheet,args_t.output_regmdl_file)
